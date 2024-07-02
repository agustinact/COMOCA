from tkinter import Toplevel
from tkinter.messagebox import askokcancel
from openpyxl import Workbook, load_workbook
from Auxiliares.exc2pdf import excel_to_pdf
from openpyxl.utils import get_column_letter
import openpyxl
import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from Auxiliares.Send_mail import enviar_correo_con_adjuntos

def adjudicar_remito(n_remito, tupla_cliente, lista_datos, factura):
    print("tupla cliente: ", tupla_cliente )
    wb = load_workbook('remitos/remito_plantilla.xlsx')
    excel_file = f'remitos/C{n_remito}.xlsx'
    sheet_name = 'BASE'    
    pdf_file   = f'remitos/C{n_remito}.pdf'
    archivos_adjuntos = []
    hoja = wb.worksheets[0]
    # numero remito
    hoja['I3'] = n_remito
    # fecha
    hoja['I5'] = datetime.date.today()
    # nombre cliente
    hoja['C8'] = tupla_cliente[1]
    # domicilio
    hoja['C9'] = tupla_cliente[3]
    # cuit
    hoja['C10'] = tupla_cliente[4]
    # mail
    hoja['C11'] = tupla_cliente[5]
    # telefono
    hoja['H8'] = tupla_cliente[6]
    # CP
    hoja['H9'] = tupla_cliente[7]
    # provincia
    hoja['H10'] = tupla_cliente[8]

    indice_fila    = 14
    suma_bultos = 0
    suma_pesos  = 0

    for item in lista_datos:
        lista_pesos = []
        total_peso = sum(item[4])
        for peso in item[4]:
            lista_pesos.append(peso)
        datos_fila =[item[0], f"{item[1]} - Tropa: {item[2]}. {lista_pesos}", item[3], total_peso]
        
        insertar_fila_y_ajustar_alto(hoja, indice_fila, datos_fila)
        suma_bultos += item[3]
        suma_pesos += total_peso
        indice_fila += 1
        
    hoja[f'H{indice_fila}'] = suma_bultos
    hoja[f'I{indice_fila}'] = suma_pesos
    
    # Guardar archivo excel
    wb.save(excel_file)
    archivos_adjuntos.append(excel_file)
    # Enviar email
    
    send_mail = askokcancel(f"Comprobante {n_remito}", f"Desea enviar el comprobante de salida al Cliente: {tupla_cliente[1]} ")
    if send_mail:
        enviar_correo_con_adjuntos(tupla_cliente[5], n_remito, tupla_cliente[1], factura, archivos_adjuntos)
    
    
def insertar_fila_y_ajustar_alto(hoja, indice_fila, datos_fila):
    
    # Insertar una nueva fila en la posición especificada
    hoja.insert_rows(indice_fila)
    hoja.merge_cells(f'B{indice_fila}:G{indice_fila}')
    celda = hoja[f'B{indice_fila}']
    celda.alignment = openpyxl.styles.Alignment(wrap_text=True)  # Ajustar el texto
    texto = datos_fila[1]

    # Insertar los datos en la nueva fila
    #Articulo
    hoja[f'A{indice_fila}'] = datos_fila[0]
    #Detalle
    hoja[f'B{indice_fila}'] = texto
    #Bultos
    hoja[f'H{indice_fila}'] = datos_fila[2]
    #Pesos
    hoja[f'I{indice_fila}'] = datos_fila[3]
    
    alto_texto = max([len(line) for line in texto.split('\n')])  # Obtener la longitud máxima de las líneas
     # Establecer la altura de la fila en función de la altura necesaria para el texto
    alto_fila = int(alto_texto / 2) + 1  # Puedes ajustar este factor para obtener una mejor visualización
    hoja.row_dimensions[indice_fila].height = alto_fila
    